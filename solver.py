import numpy as np
import pandas as pd
import pulp
from pulp import *
from openpyxl import Workbook

IN_FILE = "Example_data/Garage_Data.xlsx"

# ---- Read data ----
df = pd.read_excel(IN_FILE, sheet_name="Sheet1")
df = df.dropna(subset=[df.columns[0]])
names = df["Garage DMUs"].tolist()

input_cols  = [c for c in df.columns if c.strip().endswith("(I)")]
output_cols = [c for c in df.columns if c.strip().endswith("(O)")]

X = df[input_cols].to_numpy(float)      # Input var. list
Y = df[output_cols].to_numpy(float)     # Output var. list

EPS = 1e-9


def dea_solver(X, Y, orient: Literal["io","oo"] = "oo"):

    n, m = X.shape  # n: total DMUs m: input var. num.
    s = Y.shape[1]  # s: output var. num.

    eff_rows, lam_rows, slack_rows = [], [], []

    if orient not in ["io", "oo"]:
        raise ValueError("orientation must be either 'io' or 'oo'")

    elif orient == "io":

        for o in range(n):
            # --- Minimize theta (efficiency score), CRS ---
            p1 = LpProblem("io_crs", LpMinimize)
            theta = LpVariable("theta", lowBound=0)
            lam = [LpVariable(f"l{j}", lowBound=0) for j in range(n)]
            p1 += theta
            for i in range(m):
                p1 += lpSum(lam[j] * X[j, i] for j in range(n)) <= theta * X[o, i], f"in{i}"
            for r in range(s):
                p1 += lpSum(lam[j] * Y[j, r] for j in range(n)) >= Y[o, r], f"out{r}"
            p1.solve(GUROBI())
            theta_star = value(theta)

            dualX = [abs(p1.constraints[f"in{i}"].pi) for i in range(m)]
            dualY = [abs(p1.constraints[f"out{r}"].pi) for r in range(s)]

            # --- Fix theta, maximize total slacks ---
            p2 = LpProblem("iocrs_slack", LpMaximize)
            lam2 = [LpVariable(f"l{j}", lowBound=0) for j in range(n)]
            sX = [LpVariable(f"sx{i}", lowBound=0) for i in range(m)]
            sY = [LpVariable(f"sy{r}", lowBound=0) for r in range(s)]
            p2 += lpSum(sX) + lpSum(sY)
            for i in range(m):
                p2 += lpSum(lam2[j] * X[j, i] for j in range(n)) + sX[i] == theta_star * X[o, i]
            for r in range(s):
                p2 += lpSum(lam2[j] * Y[j, r] for j in range(n)) - sY[r] == Y[o, r]
            p2.solve(GUROBI())

            lam_val = [v.varValue or 0.0 for v in lam2]
            sX_val = [v.varValue or 0.0 for v in sX]
            sY_val = [v.varValue or 0.0 for v in sY]

            Xeff = [theta_star * X[o, i] - sX_val[i] for i in range(m)]
            Yeff = [Y[o, r] + sY_val[r] for r in range(s)]

            phi = 1.0 / theta_star if theta_star else 0.0
            eff_rows.append([o + 1, theta_star, phi] + dualX + dualY)
            lam_rows.append([o + 1] + lam_val)
            slack_rows.append([o + 1] + sX_val + sY_val + [o + 1] + Xeff + Yeff)


    elif orient == "oo":

        for o in range(n):
            # --- Maximize phi (output expansion rate), CRS---
            p1 = LpProblem("oo_crs_dual", LpMaximize)
            phi = LpVariable("phi", lowBound=0)
            lam = [LpVariable(f"lambda{j}", lowBound=0) for j in range(n)]
            p1 += phi
            inC, outC = [], []
            for i in range(m):
                c = lpSum(lam[j]*X[j, i] for j in range(n)) <= X[o, i]
                p1 += c, f"in{i}"; inC.append(f"in{i}")
            for r in range(s):
                c = lpSum(lam[j]*Y[j, r] for j in range(n)) >= phi*Y[o, r]
                p1 += c, f"out{r}"; outC.append(f"out{r}")

            p1.solve(GUROBI())

            phi_star = value(phi)

            dualX = [abs(p1.constraints[f"in{i}"].pi) for i in range(m)]
            dualY = [abs(p1.constraints[f"out{r}"].pi) for r in range(s)]

            # --- Fix phi, maximize total slacks ---
            p2 = LpProblem("oo_crs_slack", LpMaximize)
            lam2 = [LpVariable(f"l{j}", lowBound=0) for j in range(n)]
            sX = [LpVariable(f"sx{i}", lowBound=0) for i in range(m)]
            sY = [LpVariable(f"sy{r}", lowBound=0) for r in range(s)]
            p2 += lpSum(sX) + lpSum(sY)
            for i in range(m):
                p2 += lpSum(lam2[j]*X[j, i] for j in range(n)) + sX[i] == X[o, i]
            for r in range(s):
                p2 += lpSum(lam2[j]*Y[j, r] for j in range(n)) - sY[r] == phi_star*Y[o, r]

            p2.solve(GUROBI())

            lam_val = [v.varValue or 0.0 for v in lam2]
            sX_val  = [v.varValue or 0.0 for v in sX]
            sY_val  = [v.varValue or 0.0 for v in sY]

            Xeff = [X[o, i] - sX_val[i] for i in range(m)]
            Yeff = [phi_star*Y[o, r] + sY_val[r] for r in range(s)]

            theta = 1.0/phi_star if phi_star else 0.0
            eff_rows.append([o+1, theta, phi_star] + dualX + dualY)
            lam_rows.append([o+1] + lam_val)
            slack_rows.append([o+1] + sX_val + sY_val + [o+1] + Xeff + Yeff)

    return orient, eff_rows, lam_rows, slack_rows, n, m, s

orient, eff_rows, lam_rows, slack_rows, n, m, s = dea_solver(X, Y, orient="oo")
# ---- Write Excel ----
RES_FILE = f"Results/{orient}crs_results.xlsx"

wb = Workbook()

ws1 = wb.active; ws1.title = "EffWeights"
ws1.append(["DMU", "Theta", "Phi"] +
            [f"dualX_{i+1}" for i in range(m)] +
            [f"dualY_{r+1}" for r in range(s)])
for row in eff_rows: ws1.append(row)

ws2 = wb.create_sheet("Lambda")
ws2.append(["DMU"] + [f"lambda_{j+1:>2}" for j in range(n)])
for row in lam_rows: ws2.append(row)

ws3 = wb.create_sheet("Slacks")
ws3.append(["DMU"] + [f"slackX_{i+1}" for i in range(m)] +
        [f"slackY_{r+1}" for r in range(s)] +
        ["DMU"] + [f"Xeff_{i+1}" for i in range(m)] +
        [f"Yeff_{r+1}" for r in range(s)])
for row in slack_rows: ws3.append(row)

wb.save(RES_FILE)

print(f"{'DMU':<4}{'Name':<14}{'Theta':>10}{'Phi':>10}  Eff")
for i, row in enumerate(eff_rows):
    eff = "YES" if abs(row[1]-1) < 1e-6 else ""
    print(f"{row[0]:<4}{names[i][:13]:<14}{row[1]:>10.5f}{row[2]:>10.5f}  {eff}")
print(f"\nEfficient DMUs: {sum(1 for r in eff_rows if abs(r[1]-1)<1e-6)}/{n}")
print("Saved:", RES_FILE)

